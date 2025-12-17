import difflib
import base64

import json
import re
import requests
import pandas as pd
import sys
import os
import time
from openpyxl import load_workbook
import datetime
import time

def sale_order(data):
    # 打印输入的字符串内容，方便调试
    print("解析字符串："+data)
    
    # 公司简称与全称对应字典
    match_dict =  {
        '友升':'友升启繁信息',
        '图灵':"浙江网新图灵电子有限公司",
        "辰晔":"上海辰晔信息科技有限公司",
        "山东伟联":"山东伟联科贸有限公司-客户",
        "海德":"上海海德众业信息科技有限公司",
        "金陵":"上海金陵网络科技有限公司",
        "沈阳骏腾":"沈阳神州骏腾信息技术有限公司",
        "中天":"广州中天鼎力科技有限公司",
    }

    # 定义一组正则表达式，用于匹配并删除文本中一些格式化内容，如括号中含+号内容、逗号、价格字样、时间格式等
    format_re_pattern = [
        '\（[^\）]+(?=\+)[^\）]+\）',   # 中文括号内含+号的内容
        '\([^\)]+(?=\+)[^\)]+\)',     # 英文括号内含+号的内容

        '，',                        # 中文逗号
        ',',                         # 英文逗号
        '价格',                      # 文字“价格”

        '\（[^\）]+(?=SEC)[^\）]+\）',  # 中文括号内含SEC的内容
        '\([^\)]+(?=SEC)[^\)]+\)',      # 英文括号内含SEC的内容

        '[（(][^)）]*\d{4}年[-\s]?\d{1,2}月[^)）]*[)）]',  # 括号内的年月信息，格式如“2023年 5月”
        '\([^)]*\d{4}年[-\s]?\d{1,2}月[^)]*\)',           # 同上英文括号
    ]
    
    # 遍历这些正则表达式，查找匹配并打印匹配内容，然后将其替换为空格
    for i in format_re_pattern:
        if re.search(rf'{i}',data):
            print("执行："+i+" 匹配到的字符串为："+ (re.search(rf'{i}',data)).group(0))    
        data = re.sub(rf"{i}"," ",data)
    
    # 格式化后的数据输出
    print("格式化后数据为 "+data)

    res_list = []  # 最终结果列表，存放每个产品的字典信息
    
    # 用于匹配客户公司名称的两种句式
    company_repattern = [
        "给(.*)做个",
        "给(.*)出库",
        "给(.*)做出库"

    ]
    
    # 尝试从字符串中提取公司名称，并匹配字典中的全称
    for i in company_repattern:
        company = re.search(rf'{i}',data)
        if company:
            company_name = company.group(1)
            for k in match_dict:
                if company.group(1) == k:
                    company_name = match_dict[k]
            break  # 找到公司后停止查找
    
    # 判断是否包含“未税”关键词，决定税务状态
    is_tax = re.search(r'未税',data)

    # 物流速度默认“标快”，如果匹配到“隔天”或“当天”则更新
    freight_speed_list = ['隔天','当天']
    freight_speed = "标快"
    for i in freight_speed_list:
        tmp = re.search(rf'{i}',data)
        if tmp:
            freight_speed = tmp.group(0)
    
    # 销售员名单，默认为“徐志康”，匹配到则更新
    saler = "徐志康"
    saler_list=['徐志康','袁秋璟','於罡','李平','何兰军','王武胜','杨洪星'] 
    for i in saler_list:
        tmp = re.search(rf'{i}',data)
        if tmp:
            saler = tmp.group(0)
    
    # 物流公司名单，默认物流联系人为“王师傅”，匹配到公司名称则替换
    freight_company_list = ["顺丰","京东","中通","圆通","物流"]
    freight = "王师傅"
    for i in freight_company_list:
        if re.search(rf'{i}',data):
            freight = i

    # 判断打印单据责任人：配施、发艳或无
    if "配施" in data:
        freight_print = "配施"
    elif "发艳" in data:
        freight_print = "发艳"
    else:
        freight_print = "无"
    
    # 根据是否找到“未税”决定税务状态字符串
    if is_tax == None:
        tax = "否"
    else:
        tax = "未税"
    
    product_list = []

    # 清理数据中“配施”和“发艳”字样，防止干扰后续解析
    data = data.replace('配施','').replace('发艳','')
    
    # 尝试从字符串中提取客户地址信息，格式：地址：xxx
    customer_info = re.search(r'地址：(.*)',data)
    customer_info_list = None
    if customer_info:
        customer_info = customer_info.group(1)
        print(customer_info)
        
        # 按空格拆分客户信息（姓名、电话、地址）
        customer_info_list = customer_info.split(" ")
        customer_info_list = [i for i in customer_info_list if i]  # 过滤空字符串
        
        count = 0
        # 根据内容长度推断电话、姓名和地址
        for i in customer_info_list:
            if len(i) == 11:  # 11位数字认为是电话
                customer_phonenumber = customer_info_list[count]
            elif len(i) <= 4:  # 短字符串认为是姓名
                customer_name = customer_info_list[count]
            else:  # 其余认为是地址
                customer_address = customer_info_list[count]
            count = count + 1
        
        # 拼接成字符串形式，方便后续存储
        customer_info_list = customer_name +" "+customer_phonenumber+" "+customer_address

    # 判断是否有多个设备，依据“元...+”的模式提取多个设备信息
    if re.search(r"元(.*?)\+",data):
        res = re.findall(r"元(.*?)\+",data)
        print("解析到多个设备")
        print(res)
        
        # 当多个设备时，切分并分别存入列表
        if len(res)>1:
            for i in res:
                print("当前数据为："+data)
                resss = re.search(r'出库 ?：(.*)',data).group(1).split(i)[0]+i
                product_list.append(resss)
                data = data.replace(resss,"")
        else:
            print("包含2个设备")
            i = res[0]
            product_list = re.search(r'出库 ?：(.*)',data).group(1).split(i+"+")
            product_list[0] = product_list[0]+i
    else:
        # 没有“元...+”格式，认为只有一个设备，匹配“出库：”后的产品信息
        print("解析到一个设备")
        product_list.append(re.search(r'出库 *?：(.*台|.*个|.*块|.*片|.*套|.*根|.*瓶|.*条)',data).group(1))
    
    print("解析后列表为：："+str(product_list))
    
    # 从产品列表字符串中提取备注内容，提取最后一个“台/个/块/片/套/根”后面的文本作为备注
    remark_tmp = re.search(r'(?:台|个|块|片|套|根)(?!.*(?:台|个|块|片|套|根|条)).*?(.+)$', data)
    if remark_tmp:
        remark = remark_tmp.group(1)
    else:
        remark = ""
    # 遍历每个产品信息，使用正则提取产品名称、价格、数量和单位
    for i in product_list:
        print(i)
        res = re.search(r"(.*?) ?(\d+) ?元 ?\*? ?(\d+) ?(台|个|块|套|根)?",i)
        
        # print(res.group(0))
        print(res)
        if res == None:
            res = re.search(r"(.*?) ?(\d+) ?元",i)
            number = "1"   # 数量
        else:
            number = res.group(3)   # 数量
        price = res.group(2)    # 单价
        product_name = res.group(1)  # 产品名称
        
        # 构造字典，追加到结果列表中
        res_list.append({
            "主机名称":product_name,
            "数量":number,
            "价格":price,
            "物流公司":freight,
            "物流速度":freight_speed,
            "动作":"出库",
            "公司名字":company_name,
            "未税":tax,
            "销售员":saler,
            "客户地址":customer_info_list,
            "打印单据" : freight_print,
            "备注" : remark
        })
    
    # 打印最终结果
    print(res_list)
    return res_list

def print_freight_order(sale_order_list):
    # 定义销售员姓名，用于编号
    saler_names = ["何兰军", "徐志康", "袁秋璟", "於罡", "王武胜", "李平", "杨洪星","吕鑫","赵浩然"]

    # 从销售订单中获取打印模板名称
    sheet_name = sale_order_list[0]["打印单据"]
    file_path = 'C://kpfile//物流单据模板-' + sheet_name + '.xlsx'

    # 加载 Excel 文件
    wb = load_workbook(file_path)

    # 判断 sheet 是否存在
    if sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
    else:
        print(f"工作表 '{sheet_name}' 不存在！")
        exit()

    product_name = ""
    product_list = []

    # 构造打印的产品信息字符串
    for i in sale_order_list:
        # 拆装分析，传入主机信息（run_chaizhuang 是自定义的拆解函数）
        tmp = run_chaizhuang("母机：" + i['主机名称'] + " " + i["数量"] + "个" + "拆：", analyze=True)

        # 判断单位类型
        if tmp[1] == "Mixed":
            d_type = '台'
        elif tmp[1] == 'memory':
            d_type = '根'
        else:
            d_type = '个'

        # 根据不同模板格式生成产品信息字符串
        if sheet_name == "配施":
            product_list.append(product_name + tmp[0].split('+')[0] + "*" + i["数量"])
        else:
            product_list.append(product_name + tmp[0].replace("+", "/") + " 数量：" + i["数量"] + d_type)

    # 打印输出生成的字符串列表
    print("写入字符串为：", product_list)

    # 多个产品名之间加换行符
    for i in range(0, len(product_list) - 1):
        product_list[i] = product_list[i] + '\n'

    # 拼接成最终要写入表格的字符串
    for i in product_list:
        product_name = product_name + i

    # 拆分客户地址（假设为“省 市 地址”）
    customer_info_list = sale_order_list[0]['客户地址'].split(" ")

    # 配施模板的写入逻辑
    if sheet_name == "配施":
        ws['A3'] = sale_order_list[0]['销售员']              # 销售员姓名
        ws['A5'] = product_name                              # 产品信息
        ws['E3'] = customer_info_list[0]                     # 省
        ws['E4'] = customer_info_list[1]                     # 市
        ws['E5'] = customer_info_list[2]                     # 地址
        ws['E8'] = time.strftime("%Y/%m/%d", time.localtime())  # 当前日期

    # 发艳模板的写入逻辑
    elif sheet_name == "发艳":
        # 读取/更新编号计数器（防止重复编号）
        counter_file = 'C:\\kpfile\\wf_counter.json'
        today = time.strftime('%Y%m%d', time.localtime())
        count = 1
        saler = str(saler_names.index(sale_order_list[0]['销售员']) + 1)  # 销售员编号

        # 读取计数文件
        if os.path.exists(counter_file):
            with open(counter_file, 'r') as f:
                data = json.load(f)
                if data.get('date') == today:
                    count = data.get('count', 0) + 1

        # 更新计数文件
        with open(counter_file, 'w') as f:
            json.dump({'date': today, 'count': count}, f)

        # 生成唯一编号，如 WF-20250521-3-2
        id_str = f"WF-{today}-{count}-{saler}"

        # 写入发艳模板内容
        ws['B6'] = product_name                              # 产品信息
        ws['B7'] = customer_info_list[2]                     # 地址
        ws['B3'] = time.strftime("%Y/%m/%d", time.localtime())  # 日期
        ws['F5'] = customer_info_list[0] + customer_info_list[1]  # 省市
        ws['E3'] = id_str                                    # 唯一编号

    # 保存 Excel 文件
    wb.save(file_path)

def chaizhuang(data):
    flag = ''  # 标记解析模式，manual（手动模式）或auto（自动模式）
    res_dict = {"入库设备":{}, "出库设备":{}}  # 用于保存入库和出库设备信息的字典
    ruku_list = []  # 入库设备列表
    chuku_list = []  # 出库设备列表

    # 先对数据做一些简单的替换，去除“价格”和中文逗号，方便后续正则匹配
    data = data.replace("价格"," ").replace("，"," ")
    
    # 尝试用手动模式匹配“出库：”部分的设备信息
    try:
        # 正则表达式匹配格式如“出库：设备名称 单价 元 * 数量 台”
        chuku_re = re.compile(r'(出库：|出库:)(.*?)(\d+)元\s?\*\s?(\d)+台')
        chuku = re.search(chuku_re,data)
        # 匹配成功则将出库设备信息添加到入库列表（这里可能是业务逻辑中的特殊情况）
        ruku_list.append({"名称": chuku.group(2), "单价": chuku.group(3), "数量": chuku.group(4)})
        flag = "manual"  # 标记为手动模式
    except:
        flag = "auto"  # 如果匹配失败，标记为自动模式

    # 匹配拆卸设备（拆：或拆:后面的内容）
    chai_device = re.search(r'(拆：|拆:)(.*)\n?', data)
    if chai_device:
        print("捕获的拆卸设备字符串为：", chai_device.group(0))
        # 找出拆卸设备中设备名称、单价、数量的三元组
        # chai_tuple = re.findall(r'(.*?)(\d+)\s?元\*?(\d+)?\+?', chai_device.group(2))
        chai_tuple = re.findall(r'([\w\W]*?)\s*(\d+)\s*元\*?(\d+)?(台|个)?\s*\+?', chai_device.group(2))

        for chuku_device in chai_tuple:
            # 如果数量为空，则默认数量为1
            if chuku_device[2] == "":
                chuku_num = "1"
            else:
                chuku_num = chuku_device[2]
            # 将拆卸设备添加到入库设备列表
            ruku_list.append({"名称": chuku_device[0], "数量": chuku_num, "单价": chuku_device[1]})

    # 匹配母机信息，正则中利用了非贪婪匹配和断言，匹配母机名称直到下一个关键词或设备数量等
    mj_re = re.compile(r'(母机：|母机:)(.*?)(?=\s*?\*?\d+(台|个)|拆：|拆:|加：|加:|\n)', re.DOTALL)
    mj = re.search(mj_re, data)
    if mj == None:
        print("未匹配到母机没有数量，默认为 1 ")
        # 如果没有匹配到数量，尝试匹配母机名称部分
        mj_re = re.compile(r'(母机：|母机:)(.*?)(?=(拆：|拆:|加：|加:))')
        mj = re.search(mj_re, data)
        mj_num = "1"
    else:
        # 尝试提取母机数量
        mj_num = re.search(r'(母机：|母机:).*?\s(\d+)(台|个)', data)
        if mj_num:
            mj_num = mj_num.group(2)
        else:
            mj_num = '1'
    # 将母机添加到出库设备列表，单价写为“库存价格”表示未知或默认
    chuku_list.append({"名称": mj.group(2), "数量": mj_num, "单价": "库存价格"})

    # 匹配“加：”或“加:”之后到“拆:”或结尾的内容，表示增加的设备
    print(data)
    jia_tmp = re.search(r'(加:|加：)(.*?)(拆:|拆：|$)', data ,re.DOTALL)

    if jia_tmp != None:
        print("匹配到的加的设备字符串为：", jia_tmp.group(2))
        # 去掉括号中包含“到货”或“预计”的备注信息，避免干扰数量解析
        tmp = re.sub(r'[\(\（][^()\）\（）]*(到货|预计)[^()\）\（）]*[\)\）]', '', jia_tmp.group(2))+"+"
        # 匹配加设备名称和数量，如“设备名称 *数量 +”
        items = re.findall(r'(.+?)\+', tmp)
        jia_tuple = []
        for item in items:
            match = re.match(r'(.+?)\*([0-9]+)', item)
            if match:
                name, qty = match.group(1), match.group(2)
            else:
                name, qty = item, "1"
            jia_tuple.append((name.strip(), qty, ''))

        for chuku_device in jia_tuple:
            print(chuku_device)
            # 数量为空默认1，否则去除前缀的*号
            if chuku_device[1] == None or chuku_device[1] == '':
                ruku_num = "1"
            else:
                ruku_num = chuku_device[1]
            # 加入到出库设备列表
            chuku_list.append({"名称": chuku_device[0], "数量": ruku_num.replace("*", ""), "单价": "库存价格"})

    # 结果字典赋值
    res_dict["入库设备"] = ruku_list
    res_dict["出库设备"] = chuku_list
    print(res_dict)
    return [res_dict, flag]

def remove_serial(name):
    # 删除孤立的 8~12 位英数字组合（可能是裸露的序列号）
    name = re.sub(r'\b[A-Za-z0-9]{9,12}\b', '', name)
    # 多余空格清除
    name = re.sub(r'\s+', ' ', name).strip()
    return name

def segment(text, kw):
    m = re.search(re.escape(kw), text, re.IGNORECASE)
    if not m: return ""
    l = text.rfind('/', 0, m.start())
    r = text.find('/', m.end())
    return text[l+1:r if r != -1 else len(text)].strip('/')


def create_product_info(res,analyze=False):
    # print("res     ",res)
    dismantled_device_list = [{remove_serial(i['名称']): i['数量']} for i in res['入库设备'][0:]]
    added_device_list = [{remove_serial(i['名称']): i['数量']} for i in res['出库设备'][1:]]
    excel_file_path = 'C://kpfile//货物简称.xlsx'
    df = pd.read_excel(excel_file_path)

    # 获取列名
    brand_column = df.columns[0]        # 品牌
    sub_brand_column = df.columns[1]    # 子品牌
    name_column = df.columns[2]         # 名称（匹配字段）
    type_column = df.columns[3]         # 类型
    unit_column = df.columns[4]         # 单位

    # 获取设备名称（标准化处理）
    mothermachine_name = res['出库设备'][0]['名称'].strip()
    mothermachine_name = remove_serial(mothermachine_name)
    mothermachine_num = res['出库设备'][0]['数量']

    matched_row = None

    print("尝试匹配以下名称项：")
    for index, row in df.iterrows():
        name_value = str(row[name_column]).strip()
        if not name_value or name_value == 'nan':
            continue
        if name_value.lower() in mothermachine_name.lower():
            print(f"[匹配✅] 名称: {name_value}, 品牌: {row[brand_column]}, 子品牌: {row[sub_brand_column]}, 类型: {row[type_column]}, 单位: {row[unit_column]}")
            matched_row = row
            break  # 匹配到后立即停止
        else:
            pass

    # 取匹配到的行
    if matched_row is not None:
        brand = matched_row[sub_brand_column]
        model = matched_row[name_column]
        best_type = matched_row[type_column]
        best_unit = matched_row[unit_column]
    else:
        brand = "Unknown"
        model = "Unknown"
        best_type = "Unknown"
        best_unit = "Unknown"

    # 读取 Excel 文件的 CPU 工作表
    file_path = "C://kpfile//device_type.xlsx"  # 你的 Excel 文件路径
    cpu_size = pd.read_excel(file_path, sheet_name="cpu")["名称"].dropna().tolist()
    raid_size = pd.read_excel(file_path, sheet_name="raid")["名称"].dropna().tolist()
    memory_size = pd.read_excel(file_path, sheet_name="memory")["名称"].dropna().tolist()
    ethernet_size = pd.read_excel(file_path, sheet_name="ethernet")["名称"].dropna().tolist()
    hdd_size = pd.read_excel(file_path, sheet_name="HDD")["名称"].dropna().tolist()
    ssd_size = pd.read_excel(file_path, sheet_name="SSD")["名称"].dropna().tolist()
    battery_size = pd.read_excel(file_path, sheet_name="battery")["名称"].dropna().tolist()
    gcard_size = pd.read_excel(file_path, sheet_name="显卡")["名称"].dropna().tolist()
    monitor_size = pd.read_excel(file_path, sheet_name="monitor")["名称"].dropna().tolist()


    accessory = [{"RAID":raid_size},{"网卡":ethernet_size},{"CPU":cpu_size},
                {"电源":battery_size},{"显卡":gcard_size},{"内存":memory_size},{"HDD":hdd_size},{"SSD":ssd_size},{"显示器":monitor_size}]

    mothermachine_accessory = {key:[] for item in accessory for key in item.keys()}
    dismantled_device_accessory = {key: [] for item in accessory for key in item.keys()}
    added_device_accessory = {key: [] for item in accessory for key in item.keys()}

    cpu_recompile = [re.compile(r'\b\d{3,4}[A-Z]{1,2}\b'),
                    re.compile(r'\bi[0-9]\-\d{3,}([A-Z]{0,2})\b'),
                    re.compile(r'\b[A-Z0-9]{4,}\b'),
                    ]

    print("=====================母机信息=====================")
    print(mothermachine_name)
    print("==================正在匹配母机信息=================")
    for item in accessory:
        for key, values in item.items():
            flag = False
            for value in values:
                value = str(value)
                if key == "CPU":
                    for i in cpu_recompile:
                        num1 = re.search(i, value)
                        if num1:
                            if num1.group(0).upper() in mothermachine_name.upper():
                                print("匹配成功: "+num1.group(0)+"\t"+repr(value)+"语句为："+str(i),end="")                            
                                num_pattern = r"/(\d)\*{}".format(num1.group(0))
                                print("匹配语句为："+str(num_pattern))
                                num2 = re.search(num_pattern, mothermachine_name)
                                if num2:
                                    mothermachine_accessory[key].append({num1.group(0):num2.group(1)})
                                else:
                                    mothermachine_accessory[key].append({num1.group(0):"1"})
                                flag=True
                                break
                else:
                    try:
                        v_comp = re.escape(value)
                    except:
                        v_comp = rf'{value}'                    
                    mothermachine_name = mothermachine_name + '/'
                    num1 = re.search(re.escape(v_comp.upper()), mothermachine_name.upper())
                    if num1:
                        print("num1:",num1.group(0),v_comp)
                        flag=True
                        print("匹配成功: "+v_comp+"\t"+key,end="\t")
                        num_pattern = r'/[^/]*?{}[^/]*?\*(\d)[^\d/]*?/'.format(num1.group(0))
                        num_pattern2 = r"(\d)+\*{}".format(num1.group(0))
                        print(num_pattern,"\t",num_pattern2)
                        num2 = re.search(num_pattern, mothermachine_name)
                        num3 = re.search(num_pattern2, mothermachine_name)
                        # pstrpattern = rf'[^/]*{re.escape(value)}[^/]*'
                        pstr = segment(mothermachine_name,value)
                        
                        if "SSD" in pstr or "固态" in pstr or "M.2" in pstr:
                            if key != "SSD":
                                break
                        elif "HDD" in pstr or '3.5' in pstr:
                            if key != "HDD":
                                break
                        if num2:
                            mothermachine_accessory[key].append({num1.group(0):num2.group(1)})
                            print("数量匹配为："+num2.group(1)+"\t表达式为："+num_pattern)
                        elif num3:
                            mothermachine_accessory[key].append({num1.group(0):num3.group(1)})
                            print("数量匹配为："+num3.group(1)+"\t表达式为："+num_pattern2)
                        else:
                            mothermachine_accessory[key].append({num1.group(0):"1"})
                            print("未搜多到数量，默认为：1")
                        break
                if flag:
                    break   

    if len(dismantled_device_list) > 0:
        print("=======匹配拆卸的设备信息=======")
        print("拆卸的设备列表:",end="")
        print(dismantled_device_list)
        for device in dismantled_device_list:
            cpu_found = False  # 标记是否找到 CPU，避免重复匹配
            for item in accessory:
                for key, values in item.items():
                    for value in values:
                        value = str(value)
                        try:
                            v_comp = re.escape(value.upper())
                        except:
                            v_comp = rf'{value}'
                        if key == "CPU":
                            for i in cpu_recompile:
                                num1 = re.search(i, value)
                                if num1 and num1.group(0) in list(device.keys())[0]:
                                    dismantled_device_accessory[key].append({num1.group(0):list(device.values())[0]})
                                    print("匹配成功： "+list(device.keys())[0]+"\t"+num1.group(0))
                                    cpu_found = True
                                    break  # 退出 CPU 正则匹配
                                if cpu_found:
                                    break  # 退出 CPU 处理
                        else:
                            d_name = list(device.keys())[0]
                            if re.search(v_comp, d_name.upper()):
                                if "SSD" in d_name or "固态" in d_name or "M.2" in d_name:
                                    if key != "SSD":
                                        break
                                elif "HDD" in d_name or '3.5' in d_name:
                                    if key != "HDD":
                                        break
                                if key in mothermachine_accessory:
                                    print("匹配成功： "+list(device.keys())[0]+"\t"+value)
                                    dismantled_device_accessory[key].append({value:list(device.values())[0]})
                                    break

    print("=======匹配加装的设备信息=======")
    print(added_device_list)
    for device in added_device_list:
        cpu_found = False  # 标记是否找到 CPU，避免重复匹配
        for item in accessory:
            for key, values in item.items():
                for value in values:
                    value = str(value)      
                    try:
                        v_comp = re.escape(value.upper())
                    except:
                        v_comp = rf'{value}'
                    if key == "CPU":
                        for i in cpu_recompile:
                            num1 = re.search(i, value)
                            if num1 and num1.group(0) in list(device.keys())[0]:
                                added_device_accessory[key].append({num1.group(0):list(device.values())[0]})
                                print("匹配成功： "+list(device.keys())[0]+"\t"+num1.group(0))
                                cpu_found = True
                                break  # 退出 CPU 正则匹配
                            if cpu_found:
                                break  # 退出 CPU 处理
                    else:
                        for value in values:
                            d_name = list(device.keys())[0]
                            d_name = re.sub(r'\（[^()]*\）','',d_name)
                            d_name = re.sub(r'\([^()]*\)','',d_name)
                            try:
                                v_comp = re.escape(value)
                            except:
                                v_comp = rf'{value}'
                            num1 = re.search(v_comp, d_name)
                            if num1:
                                if "SSD" in d_name or "固态" in d_name or "M.2" in d_name :
                                    if key != "SSD":
                                        break
                                elif "HDD" in d_name or '3.5' in d_name:
                                    if key != "HDD":
                                        break
                                print("匹配成功: "+str(value)+"\t"+d_name+"类型为：",key)
                                added_device_accessory[key].append({num1.group(0):list(device.values())[0]})
                                cpu_found = True
                                break
                    if cpu_found:
                        break  # 完全终止 CPU 匹配，防止重复      
            if cpu_found:
                break  # 完全终止 CPU 匹配，防止重复

    print(mothermachine_accessory)
    print(added_device_accessory)
    print(dismantled_device_accessory)
    
    # 修正母机配件中数量格式为 "2x" / "2X" / "2*x" 的情况
    for device_type, specs in mothermachine_accessory.items():
        for spec_dict in specs:
            for spec in list(spec_dict.keys()):
                pattern = re.compile(r'(\d+)\s*[xX＊*]\s*' + re.escape(spec), re.IGNORECASE)
                match = pattern.search(mothermachine_name)
                if match:
                    qty = match.group(1)
                    old_qty = spec_dict[spec]
                    if qty != old_qty:
                        spec_dict[spec] = qty
                        print(f"[✔ 修正数量] 类型: {device_type}, 型号: {spec}, 数量: {old_qty} → {qty}")
    print("修改后的母机设备为：",end="")
    print(mothermachine_accessory)
    mother = mothermachine_accessory
    added = added_device_accessory
    removed = dismantled_device_accessory
    if mother['电源']==[]:
        mother_battery_flag = False
    else:
        mother_battery_flag = True

    ### **第一步：先处理拆除设备**
    for equipment_type, specs in list(mother.items()):
        if isinstance(specs, list):  # 确保 specs 是列表
            for spec_dict in specs:  # 遍历列表中的字典
                for spec, quantity in list(spec_dict.items()):  # 访问字典的键值对
                    if spec is None or quantity is None:  # 避免 None 影响计算
                        continue
                    if equipment_type in removed:
                        for rem_dict in removed[equipment_type]:
                            if spec in rem_dict:
                                # 计算新数量
                                new_quantity = int(quantity) - int(rem_dict[spec])
                                if new_quantity > 0:
                                    spec_dict[spec] = str(new_quantity)
                                else:
                                    # 如果数量为 0 或负数，删除该规格
                                    del spec_dict[spec]
                                    if not spec_dict:  # 如果字典为空，从列表中删除
                                        specs.remove(spec_dict)
                                    break

    ### **第二步：处理新增设备**
    for equipment_type, specs in added.items():
        if isinstance(specs, list):  # 确保是列表
            for spec_dict in specs:  # 遍历新增设备
                for spec, quantity in spec_dict.items():
                    if spec is None or quantity is None:
                        continue
                    if equipment_type not in mother:
                        mother[equipment_type] = []
                    found = False
                    for existing_dict in mother[equipment_type]:
                        if spec in existing_dict:
                            # 确保数量是整数
                            if existing_dict[spec].isdigit() and quantity.isdigit():
                                existing_dict[spec] = str(int(existing_dict[spec]) + int(quantity))
                            found = True
                            break
                    if not found:
                        mother[equipment_type].append({spec: quantity})

    ### **第三步：构造输出字符串**
    result = f"{brand} {model}/"
    order = ['CPU' ,'内存', 'HDD', 'SSD','显卡', '电源', 'RAID', '网卡',"显示器"]
    
    for equipment_type in order:
        if equipment_type in mother:
            valid_specs = []
            for spec in mother[equipment_type]:  
                if isinstance(spec, dict):  
                    for k, v in spec.items():
                        if k is not None and v is not None and v != "无":
                            valid_specs.append({k: v})

            if valid_specs:
                for spec_dict in valid_specs:
                    for spec, quantity in spec_dict.items():
                        result += f"{equipment_type} {spec}*{quantity}/"
            else:
                result += f"无{equipment_type}/"

    # 如果 HDD 和 SSD 都为空，输出 "无硬盘"
    if "无HDD" in result and "无SSD" in result:
        result = result.replace("无HDD","").replace("无SSD","无硬盘")
    if "SSD " in result:
        result = result.replace("无HDD","")
    if "HDD " in result:
        result = result.replace("无SSD","")
    if "无电源" in result and mother_battery_flag == False:
        result = result.replace('无电源',"")
    result = result.replace('无RAID',"").replace("无网卡","").replace("无显卡","").replace("无显示器","").replace("**","*")
    result = re.sub(r"/{2,}","/",result).rstrip('/')  # 去掉最后的 "/"

    print(result)
    r_type = ""
    if analyze:
        if model != "Unknown":
            analyze_string = model+"+"
        else:
            analyze_string = ""
        for i in mother:
            if mother[i] == []:
                pass
            else:
                for j in mother[i]:
                    for k in j:
                        analyze_string = analyze_string+str(k+"*"+j[k]) + "+"
                        r_type = i
        plus_count = analyze_string.count('+')
        if plus_count == 1:
                return analyze_string,r_type
        return analyze_string,"Mixed"
    else:
        try:
            res = {"货品名称": result, "货品类别": best_type, "基本单位": best_unit, "产品线": "徐志康", "规格型号": brand, "数量": mothermachine_num, "品牌": "任务备货"}
        except:
            return 0
    return res

def run_chaizhuang(data,analyze=False):
    res = chaizhuang(data.replace(",","+").replace("，","+"))
    if res[1] == "auto":
        print("=====================创建品名中=====================")
        return create_product_info(res[0],analyze)
    else:
        print("==================已有品名，返回中===================")
        return res[0]

def find_best_match(target, candidates):
    best_match = None
    highest_ratio = 0.0
    best_index = -1

    for i, candidate in enumerate(candidates):
        ratio = difflib.SequenceMatcher(None, target, candidate).ratio()
        if ratio > highest_ratio:
            highest_ratio = ratio
            best_match = candidate
            best_index = i

    print(f"最佳匹配项：{best_match}")
    print(f"在列表中的索引位置：{best_index}")
    return best_match, highest_ratio, best_index

def check_or_update_date(file_path='C:\\kpfile\datetime.txt'):
    today_str = datetime.date.today().isoformat()
    if os.path.exists(file_path):
        with open(file_path, 'r') as f:
            file_date_str = f.read().strip()
        try:
            file_date = datetime.date.fromisoformat(file_date_str)
            if file_date == datetime.date.today():
                return True
        except ValueError:
            pass  # 文件内容不是有效日期，继续写入今天日期

    # 如果日期不是今天、文件不存在或内容错误，写入今天日期
    with open(file_path, 'w') as f:
        f.write(today_str)
    return False

def extract_info(data):
    def clean_segment(segment):
        # 去掉前缀中文或英文、去掉 *数字
        segment = re.sub(r'^[\u4e00-\u9fa5A-Za-z]+\s*', '', segment)
        segment = re.sub(r'\*\d+', '', segment)
        return segment.strip()

    final_res = ""

    # 提取括号中的主机型号
    match = re.search(r'[（?(]([A-Za-z0-9\-]{8})[）?)]', data)
    if match:
        final_res = match.group(1)

    # 调用外部解析函数获取货品名称
    raw_name,r_type = run_chaizhuang(data,True)  # 你需传入该函数，返回格式应为 {'货品名称': 'xxx'}

    final_res = re.sub(r'\*\d+', '', raw_name)
    return raw_name,final_res,r_type  # 去除 *后面的数字
